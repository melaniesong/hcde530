"""Excel export for consolidated UX research analysis results."""

from __future__ import annotations

import io
import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

THEMES_SUMMARY_HEADERS = ("Theme", "Severity", "Frequency (sessions)", "Summary")
QUOTES_HEADERS = ("Theme", "Quote", "Timestamp", "Source File")
SESSION_SUMMARIES_HEADERS = ("Session File", "Top Themes", "Notable Observations")
RAW_ANALYSIS_HEADERS = ("Session File", "Full JSON analysis")


def _session_themes(entry: dict[str, Any]) -> list[dict[str, Any]]:
    if "themes" in entry:
        return entry["themes"]
    if "analysis" in entry and isinstance(entry["analysis"], dict):
        return entry["analysis"].get("themes", [])
    return []


def _style_header_row(ws: Worksheet, column_count: int) -> None:
    for col in range(1, column_count + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
    ws.freeze_panes = "A2"


def _autofit_columns(ws: Worksheet, *, max_width: int = 80) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


def _write_themes_summary(ws: Worksheet, consolidated: dict[str, Any]) -> None:
    ws.title = "Themes Summary"
    for col, header in enumerate(THEMES_SUMMARY_HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)

    for row_idx, theme in enumerate(consolidated.get("themes", []), start=2):
        ws.cell(row=row_idx, column=1, value=theme.get("name", ""))
        ws.cell(row=row_idx, column=2, value=theme.get("severity", ""))
        ws.cell(row=row_idx, column=3, value=theme.get("frequency", ""))
        ws.cell(row=row_idx, column=4, value=theme.get("summary", ""))

    _style_header_row(ws, len(THEMES_SUMMARY_HEADERS))
    _autofit_columns(ws)


def _build_timestamp_lookup(
    per_session_results: list[dict[str, Any]],
) -> dict[tuple[str, str], str | None]:
    """Map (source file, quote text) to pass 1 timestamps."""
    lookup: dict[tuple[str, str], str | None] = {}
    for entry in per_session_results:
        source = entry.get("filename", "")
        for theme in _session_themes(entry):
            quote_text = theme.get("quote", "").strip()
            if quote_text:
                lookup[(source, quote_text)] = theme.get("timestamp")
    return lookup


def _resolve_quote_timestamp(
    quote: dict[str, Any],
    lookup: dict[tuple[str, str], str | None],
) -> str:
    """Prefer pass 2 timestamp; fall back to pass 1 lookup by source + quote text."""
    timestamp = quote.get("timestamp")
    if timestamp:
        return str(timestamp)

    source = quote.get("source", "")
    text = quote.get("text", "").strip()
    if not source or not text:
        return ""

    exact = lookup.get((source, text))
    if exact:
        return str(exact)

    for (lookup_source, lookup_text), lookup_ts in lookup.items():
        if lookup_source == source and lookup_ts and (
            lookup_text in text or text in lookup_text
        ):
            return str(lookup_ts)

    return ""


def _write_quotes_by_theme(
    ws: Worksheet,
    consolidated: dict[str, Any],
    per_session_results: list[dict[str, Any]],
) -> None:
    ws.title = "Quotes by Theme"
    for col, header in enumerate(QUOTES_HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)

    timestamp_lookup = _build_timestamp_lookup(per_session_results)
    row_idx = 2
    for theme in consolidated.get("themes", []):
        theme_name = theme.get("name", "")
        for quote in theme.get("quotes", []):
            ws.cell(row=row_idx, column=1, value=theme_name)
            ws.cell(row=row_idx, column=2, value=quote.get("text", ""))
            ws.cell(
                row=row_idx,
                column=3,
                value=_resolve_quote_timestamp(quote, timestamp_lookup),
            )
            ws.cell(row=row_idx, column=4, value=quote.get("source", ""))
            row_idx += 1

    _style_header_row(ws, len(QUOTES_HEADERS))
    _autofit_columns(ws)


def _write_session_summaries(
    ws: Worksheet, per_session_results: list[dict[str, Any]]
) -> None:
    ws.title = "Session Summaries"
    for col, header in enumerate(SESSION_SUMMARIES_HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)

    for row_idx, entry in enumerate(per_session_results, start=2):
        themes = _session_themes(entry)
        theme_names = [t.get("name", "") for t in themes if t.get("name")]
        observations = [t.get("summary", "") for t in themes if t.get("summary")]

        ws.cell(row=row_idx, column=1, value=entry.get("filename", ""))
        ws.cell(row=row_idx, column=2, value=", ".join(theme_names))
        ws.cell(row=row_idx, column=3, value="; ".join(observations))

    _style_header_row(ws, len(SESSION_SUMMARIES_HEADERS))
    _autofit_columns(ws)


def _write_raw_analysis(ws: Worksheet, per_session_results: list[dict[str, Any]]) -> None:
    ws.title = "Raw Analysis"
    for col, header in enumerate(RAW_ANALYSIS_HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)

    for row_idx, entry in enumerate(per_session_results, start=2):
        filename = entry.get("filename", "")
        if "analysis" in entry and isinstance(entry["analysis"], dict):
            payload = entry["analysis"]
        else:
            payload = {"themes": _session_themes(entry)}

        ws.cell(row=row_idx, column=1, value=filename)
        ws.cell(
            row=row_idx,
            column=2,
            value=json.dumps(payload, indent=2, ensure_ascii=False),
        )

    _style_header_row(ws, len(RAW_ANALYSIS_HEADERS))
    _autofit_columns(ws, max_width=120)


def export_to_excel(
    consolidated: dict[str, Any],
    per_session_results: list[dict[str, Any]],
) -> io.BytesIO:
    """
    Build a four-tab Excel workbook from pass 2 and pass 1 results.

    Parameters
    ----------
    consolidated
        Pass 2 output with a ``themes`` list (frequency, sessions, quotes).
    per_session_results
        Pass 1 outputs, each with ``filename`` and ``themes`` or ``analysis``.

    Returns
    -------
    io.BytesIO
        In-memory ``.xlsx`` file ready for Streamlit download.
    """
    workbook = Workbook()
    default_sheet = workbook.active
    if default_sheet is not None:
        workbook.remove(default_sheet)

    _write_themes_summary(workbook.create_sheet(), consolidated)
    _write_quotes_by_theme(workbook.create_sheet(), consolidated, per_session_results)
    _write_session_summaries(workbook.create_sheet(), per_session_results)
    _write_raw_analysis(workbook.create_sheet(), per_session_results)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


if __name__ == "__main__":
    from pathlib import Path

    DUMMY_CONSOLIDATED = {
        "themes": [
            {
                "name": "Promo/Discount Code Field Discoverability",
                "severity": "High",
                "frequency": 2,
                "summary": (
                    "Users consistently struggle to locate the promo or discount code "
                    "entry field during checkout."
                ),
                "sessions": ["session-1-notes.txt", "session-2-transcript.txt"],
                "quotes": [
                    {
                        "text": (
                            "Participant struggled to find the promo code field; "
                            "looked under 'Payment' first."
                        ),
                        "timestamp": None,
                        "source": "session-1-notes.txt",
                    },
                    {
                        "text": (
                            "Could not find where to enter a discount code until "
                            "facilitator hinted."
                        ),
                        "timestamp": "00:32",
                        "source": "session-2-transcript.txt",
                    },
                ],
            },
            {
                "name": "Shipping Options Visibility Below Fold",
                "severity": "High",
                "frequency": 2,
                "summary": (
                    "Critical shipping information is positioned below the visible "
                    "viewport, causing users to miss options."
                ),
                "sessions": ["session-1-notes.txt", "session-2-transcript.txt"],
                "quotes": [
                    {
                        "text": (
                            "Participant struggled to find the promo code field; "
                            "looked under 'Payment' first."
                        ),
                        "timestamp": None,
                        "source": "session-1-notes.txt",
                    },
                    {
                        "text": (
                            "Could not find where to enter a discount code until "
                            "facilitator hinted."
                        ),
                        "timestamp": "00:32",
                        "source": "session-2-transcript.txt",
                    },
                ],
            },
            {
                "name": "Positive Product Visualization and Size Information",
                "severity": "Low",
                "frequency": 2,
                "summary": (
                    "Product photos, galleries, and size guides were well-received "
                    "across sessions."
                ),
                "sessions": ["session-1-notes.txt", "session-2-transcript.txt"],
                "quotes": [
                    {
                        "text": "I didn't know I had to scroll down to see shipping options.",
                        "timestamp": None,
                        "source": "session-1-notes.txt",
                    },
                    {
                        "text": "I almost missed shipping.",
                        "timestamp": "01:05",
                        "source": "session-2-transcript.txt",
                    },
                ],
            },
        ]
    }

    DUMMY_PER_SESSION = [
        {
            "filename": "session-1-notes.txt",
            "themes": [
                {
                    "name": "Promo Code Field Discoverability",
                    "severity": "Medium",
                    "quote": (
                        "Participant struggled to find the promo code field; "
                        "looked under 'Payment' first."
                    ),
                    "timestamp": None,
                    "summary": "The promo code input is not where users expect it.",
                },
                {
                    "name": "Discount Code Field Hard to Find",
                    "severity": "High",
                    "quote": "Could not find where to enter a discount code.",
                    "timestamp": "00:32",
                    "summary": "Discount code entry required facilitator assistance.",
                },
            ],
        },
        {
            "filename": "session-2-transcript.txt",
            "themes": [
                {
                    "name": "Discount Code Field Hard to Find",
                    "severity": "High",
                    "quote": (
                        "Could not find where to enter a discount code until "
                        "facilitator hinted."
                    ),
                    "summary": "Discount code entry required facilitator assistance.",
                },
                {
                    "name": "Shipping Options Below Fold",
                    "severity": "Medium",
                    "quote": "I almost missed shipping.",
                    "summary": "Shipping options were positioned below the visible area.",
                },
            ],
        },
    ]

    output_dir = Path(__file__).resolve().parent / "output"
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / "test_export.xlsx"

    excel_bytes = export_to_excel(DUMMY_CONSOLIDATED, DUMMY_PER_SESSION)
    output_path.write_bytes(excel_bytes.getvalue())

    print(f"Wrote {output_path}")
    print(f"Size: {output_path.stat().st_size:,} bytes\n")

    from openpyxl import load_workbook

    wb = load_workbook(output_path, read_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        print(f"=== {sheet_name} ({len(rows) - 1} data rows) ===")
        print(f"Headers: {rows[0]}")
        if len(rows) > 1:
            print(f"First row: {rows[1]}")
        print()

    wb.close()
    print("Opening file in default spreadsheet app...")
    import subprocess

    subprocess.run(["open", str(output_path)], check=False)
